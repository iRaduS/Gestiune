from openpyxl import load_workbook
from openpyxl import Workbook

ticket_workbook = load_workbook('situation.xlsx')
ticket_worksheet = ticket_workbook['Sheet1']

month = input('Introdu luna (litere mici, fara prescurtari): ')
year = input('Introdu anul (numar): ')
euro = float(input('Introdu valoarea euro: '))

stats_workbook = load_workbook(month + ' ' + year + ' ' + 'franta.xlsx')
stats_worksheet = stats_workbook.active

############################################
french = {}
french["ianuarie"] = "JANVIER"
french["februarie"] = "FÉVRIER"
french["martie"] = "MARS"
french["aprilie"] = "AVRIL"
french["mai"] = "MAI"
french["iunie"] = "JUIN"
french["iulie"] = "JUILLET"
french["august"] = "AOÛT"
french["septembrie"] = "SEPTEMBRE"
french["octombrie"] = "OCTOBRE"
french["noiembrie"] = "NOVEMBRE"
french["decembrie"] = "DÉCEMBRE"
############################################

def format_worksheet(employee, curr_ws):
	curr_ws.title = employee[1] + ' ' + employee[2]
	curr_ws['B9'] = french[month] + ' ' + year
	curr_ws['B15'] = employee[1]
	curr_ws['E15'] = employee[2]
	curr_ws['B17'] = ''
	curr_ws['C19'] = employee[3]
	##########################################################
	curr_ws['C33'] = employee[14]
	curr_ws['C34'] = round(employee[14] / euro)
	##########################################################
	curr_ws['C40'] = ((employee[21] + employee[22] + employee[23]) / euro)
	curr_ws['C41'] = ((employee[25] != None) and float(employee[25]) or 0)
	# curr_ws['C42'] = round((employee[23] + employee[22] + employee[21]) / euro)
	##########################################################
	#### NEW SITUATION ###
	curr_ws['C42'] = curr_ws['C40'].value + curr_ws['C41'].value
	###########################################################
	curr_ws['C20'] = employee[6]
	if employee[7] != None:
		curr_ws['C21'] = employee[7]
	else:
		curr_ws['C21'] = 0
	if employee[8] != None:
                if isinstance(employee[8], (int, float)):
                        curr_ws['C22'] = employee[8]
                else:
                        curr_ws['C22'] = employee[3] - employee[6] - employee[7]
	else:
		curr_ws['C22'] = 0
	##########################################################
	X = curr_ws['C42'].value
	val = (curr_ws['C20'].value + 1.25 * curr_ws['C21'].value + 1.5 * curr_ws['C22'].value)
	normal = X / val
	quat = 1.25 * (X / val)
	half = 1.5 * (X / val)

	curr_ws['D20'] = normal
	curr_ws['D21'] = quat
	curr_ws['D22'] = half
	##########################################################
	curr_ws['E20'] = int(curr_ws['C20'].value * normal)
	curr_ws['E21'] = int(curr_ws['C21'].value * quat)
	curr_ws['E22'] = int(curr_ws['C22'].value * half)
	curr_ws['E19'] = curr_ws['E20'].value + curr_ws['E21'].value + curr_ws['E22'].value
	##########################################################
	curr_ws['C24'] = int(curr_ws['C20'].value * normal)
	curr_ws['C23'] = round(curr_ws['C20'].value * normal * euro)
	##########################################################
	curr_ws['C26'] = int(curr_ws['C21'].value * quat)
	curr_ws['C25'] = round(curr_ws['C21'].value * quat * euro)
	##########################################################
	curr_ws['C28'] = int(curr_ws['C22'].value * half)
	curr_ws['C27'] = round(curr_ws['C22'].value * half * euro)
	##########################################################
	curr_ws['C29'] = employee[10]
	curr_ws['C30'] = round(employee[10] / euro)
	##########################################################
	curr_ws['C31'] = employee[11] + employee[12]
	curr_ws['C32'] = round((employee[11] + employee[12]) / euro)
	##########################################################
	curr_ws['C36'] = employee[15]
	curr_ws['C37'] = employee[17]
	curr_ws['C39'] = employee[20]


############################################
for employee in stats_worksheet.iter_rows(min_row = 2, min_col = 1, values_only=True):
	curr_ws = ticket_workbook.copy_worksheet(ticket_worksheet)
	format_worksheet(employee, curr_ws)
############################################

del ticket_workbook['Sheet1']

ticket_workbook.save('fluturasi/situation ' + month + ' ' + year + '.xlsx')
