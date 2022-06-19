from openpyxl import Workbook, load_workbook
from utility import format_worksheet

try:
    ticket_workbook = load_workbook('framework_situation.xlsx')
    ticket_worksheet = ticket_workbook['framework']
except FileNotFoundError:
    print('Nu exista fisierul excel cu fluturasul model (denumit: framework_situation.xlsx si fila framework).')
    exit()

input_message = 'Introdu valoare luna (litere mici, fara prescutari), an, valoare euro (i.e. mai 2020 9.32): '
data = [str for str in input(input_message).split()]

document_name = f"{data[1]}/{data[0]} {data[1]} franta.xlsx"
try:
    stats_workbook = load_workbook(document_name)
    stats_worksheet = stats_workbook.active
except FileNotFoundError:
    print(f"Nu exista fisierul excel denumit {document_name}.")
    exit()

for employee in stats_worksheet.iter_rows(min_row=2, min_col=1, values_only=True):
    if employee[0] is not None:
        worksheet = ticket_workbook.copy_worksheet(ticket_worksheet)
        format_worksheet(employee, worksheet, data)

del ticket_workbook['framework']
ticket_workbook.save(f"output/situation {data[0]} {data[1]}.xlsx")
